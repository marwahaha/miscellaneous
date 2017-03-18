#!/usr/bin/env python
#https://docs.google.com/spreadsheets/d/1v55S9f65a4ovo58tGkV7vQUjKeuKtHJWfsFoZ8JDNPw/export?pref=2&pli=1#gid=444680994

import openpyxl
import pywapi
import smtplib
import time
import datetime
#import strptime

#from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dateutil.parser import parse as parse_date
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

#get the date
tomorrow = datetime.date.today() + datetime.timedelta(days=1)

#Announcements
New = """
<html>
    <body>
        <div align="center" style="font-size:13.3333px; text-align:center; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif" size="4">
                <span style="font-size:16pt">
                    <b>New Evening Announcements</b>
                </span>
            </font>
        </div>
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
"""
Old = """
<html>
    <body>
        <div align="center" style="font-size:13.3333px; text-align:center; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif" size="4">
                <span style="font-size:16pt">
                    <b>Old Evening Announcements</b>
                </span>
            </font>
        </div>
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
"""
NewArray = []
OldArray = []
RArray = []
Atitle = ""
Atext = ""
Alink = ""
ANew = False
Atype = "None"
Announcements = openpyxl.load_workbook('POD Evening Announcement (Responses).xlsx')
sheet = Announcements.get_sheet_by_name('Form Responses 1')
for i in range(2, sheet.get_highest_row()+1):
    add = ""
    if sheet.cell(row=i, column=1).value != None:
        Atitle = "<b>" + sheet.cell(row=i, column=2).value + "</b><br>"
        Atype = sheet.cell(row=i, column=5).value
        if sheet.cell(row=i, column=1).value.date() >= tomorrow - datetime.timedelta(days=2):
            ANew = True
        else:
            ANew = False
        if sheet.cell(row=i, column=3).value != None:
            Atext = sheet.cell(row=i, column=3).value + "<br>"
        if sheet.cell(row=i, column=4).value != None:
            Alink = "<a href=\"" + sheet.cell(row=i, column=4).value + "\">" + sheet.cell(row=i, column=4).value + "</a><br>"
        add = Atitle+Atext+Alink + """<br><br>

"""
    if ANew == True and (Atype == "Normal" or Atype == "One day only"):
        NewArray.append(add)
    elif ANew == False and Atype == "Normal" and sheet.cell(row=i, column=1).value.date() >= tomorrow - datetime.timedelta(days=5):
        OldArray.append(add)
    elif Atype == "Reoccuring":
        RArray.append(add)
    Atitle = ""
    Atext = ""
    Alink = ""
    ANew = False
    Atype = "None"
if tomorrow.strftime("%a") == "Mon":
    OldArray = RArray + OldArray
for i in range(0, len(NewArray)):
    New = New + NewArray[i]
for i in range(0, len(OldArray)):
    Old = Old + OldArray[i]
New = New +  """            </font>
        </div>
    </body>
</html>"""
Old = Old +  """            </font>
        </div>
    </body>
</html>"""
if len(NewArray) > 0:
    final = New + Old
else:
    final = Old
    
#get excel for duty and trainings
#Trainings
Trainings = openpyxl.load_workbook('Trainings.xlsx')
lastMonday = tomorrow
firstClassTraining = ""
firstClassLocation = ""
secondClassTraining = ""
secondClassLocation = ""
thirdClassTraining = ""
thirdClassLocation = ""
fourthClassTraining = ""
fourthClassLocation = ""
eveningTraining = ""
eveningLocation = ""
while lastMonday.strftime("%A") != "Monday":
    lastMonday = lastMonday - datetime.timedelta(days=1)
sheet = Trainings.get_sheet_by_name(lastMonday.strftime("%b %d").replace(" 0"," "))
for i in range(1, sheet.get_highest_column()):
    if sheet.cell(row=3, column=i).value != None:
        if tomorrow.strftime("%A").upper()[:6] == sheet.cell(row=3, column=i).value[:6]:
            firstClassTraining = sheet.cell(row=6, column=i).value
            firstClassLocation = sheet.cell(row=6, column=i+2).value
            secondClassTraining = sheet.cell(row=13, column=i).value
            secondClassLocation = sheet.cell(row=13, column=i+2).value
            thirdClassTraining = sheet.cell(row=20, column=i).value
            thirdClassLocation = sheet.cell(row=20, column=i+2).value
            fourthClassTraining = sheet.cell(row=27, column=i).value
            fourthClassLocation = sheet.cell(row=27, column=i+2).value
            eveningTraining = sheet.cell(row=34, column=i).value
            eveningLocation = sheet.cell(row=34, column=i+2).value
            break

#CHDO duty
CHDODuty = openpyxl.load_workbook('CHDODuty.xlsx')
sheet = CHDODuty.get_sheet_by_name('Sheet1')
for i in range(1, sheet.get_highest_row()):
    if sheet.cell(row=i, column=1).value != None:
        if sheet.cell(row=i, column=1).value.date() == tomorrow:
            CHDO = sheet.cell(row=i, column=2).value
for i in range(1, sheet.get_highest_row()):
    if sheet.cell(row=i, column=6).value == CHDO:
        CHDO = sheet.cell(row=i, column=5).value + " " + CHDO
            
#Cadet duty
CadetDuty = openpyxl.load_workbook('CadetDuty.xlsx')
sheet = CadetDuty.get_sheet_by_name('Spring')
for i in range(1, sheet.get_highest_row()):
    if sheet.cell(row=i, column=2).value != None:
        if sheet.cell(row=i, column=2).value.date() == tomorrow:
            RCDO = sheet.cell(row=i, column=4).value
            ACDO = sheet.cell(row=i, column=5).value
            DutyCompany = sheet.cell(row=i, column=3).value

#Upcoming Events
upcomingEvents = []
Events = openpyxl.load_workbook('Events.xlsx')
sheet = Events.get_sheet_by_name('Sheet1')
for i in range(1, sheet.get_highest_row()):
    if sheet.cell(row=i, column=1).value != None:
        if sheet.cell(row=i, column=1).value.date() >= tomorrow + datetime.timedelta(days=1) and sheet.cell(row=i, column=1).value.date() < tomorrow + datetime.timedelta(days=30):
            upcomingEvents.append(sheet.cell(row=i, column=1).value.strftime("%d%b%y") + ": " + sheet.cell(row=i, column=2).value + "<br>")

print "Getting weather..."
#get weather
weather = pywapi.get_weather_from_weather_com( "06320" , units = "imperial" )['forecasts'][1]
print "Done"

# Create message container
header = """
<html>
    <body>
        <div align="center" style="font-size:13.3333px; text-align:center; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif" size="4">
                <span style="font-size:16pt">
                    <b>CORPS OF CADETS'</b><br>
                    <b>PLAN OF THE DAY</b>
                </span>
            </font>
        </div>"""
duty = """
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
                <b>Date: """ + tomorrow.strftime("%A, %d %B %Y") + """</b><br>
                CHDO: """ + CHDO + """<br>
                RCDO: 1/c """ + RCDO.split(",")[0] + """<br>
                ACDO: 2/c """ + ACDO.split(",")[0] + """<br>
                Duty Section: """ + DutyCompany + """<br>
"""
day = ""
dayOfWeek = tomorrow.strftime("%a")
if dayOfWeek == "Mon" or dayOfWeek == "Wed":
    day = day + """                Uniform: Tropical Blue Uniform with Garrison Covers<br>
                <br>
                0600: Reveille<br>
                0615: Guardmount<br>
                0620: Morning Formation<br>
                0625: Family Style Breakfast<br>
"""
elif dayOfWeek == "Tue":
    day = day + """                Uniform: Operational Dress Uniform with Class Specific Ballcaps<br>
                <br>
                0600: Reveille<br>
                0615: Guardmount<br>
                0620: Morning Formation<br>
                0625: Family Style Breakfast<br>
"""
elif dayOfWeek == "Fri":
    day = day + """                Uniform: Tropical Blue Uniform with Combination Covers<br>
                <br>
                0600: Reveille<br>
                0615: Guardmount<br>
                0620: Morning Formation<br>
                0625: Family Style Breakfast<br>
"""
elif dayOfWeek == "Thu":
    day = day + """                Uniform: Operational Dress Uniform with Class Specific Ballcaps<br>
                <br>
                0600: Reveille<br>
                0615: Guardmount<br>
                0620: Morning Formation<br>
                0630-0810: Buffet Style Breakfast<br>
"""
elif dayOfWeek == "Sat":
    day = day + """                Uniform: Tropical Blue Uniform with Combination Covers<br>
                <br>
                0630: Reveille<br>
                0645: Guardmount<br>
                0645-0745: Buffet Style Breakfast<br>
                0745: Morning Formation<br>
"""
elif dayOfWeek == "Sun":
    day = day + """                Uniform: Tropical Blue Uniform with Combination Covers<br>
                <br>
                0715: Guardmount<br>
                0800-0915: Buffet Style Breakfast<br>
"""
done = False
if dayOfWeek != "Sat":
    if firstClassTraining != "" and firstClassTraining != None and firstClassTraining != " ":
        if firstClassTraining == secondClassTraining and secondClassTraining == thirdClassTraining and thirdClassTraining == fourthClassTraining:
            day = day + "                0700: 1/c-4/c: " + firstClassTraining
            if firstClassLocation != "" and firstClassLocation != None and firstClassLocation != " ":
                day = day + "(" + firstClassLocation + ")"
            day = day + """<br>
"""
            done = True
        else:
            day = day + "                0700: 1/c: " + firstClassTraining
            if firstClassLocation != "" and firstClassLocation != None and firstClassLocation != " ":
                day = day + "(" + firstClassLocation + ")"
            day = day + """<br>
"""
    if done == False and secondClassTraining != "" and secondClassTraining != None and secondClassTraining != " ":
        day = day + "                0700: 2/c: " + secondClassTraining
        if secondClassLocation != "" and secondClassLocation != None and secondClassLocation != " ":
            day = day + "(" + secondClassLocation + ")"
        day = day + """<br>
"""
    if done == False and thirdClassTraining != "" and thirdClassTraining != None and thirdClassTraining != " ":
        day = day + "                0700: 3/c: " + thirdClassTraining
        if thirdClassLocation != "" and thirdClassLocation != None and thirdClassLocation != " ":
            day = day + "(" + thirdClassLocation + ")"
        day = day + """<br>
"""
    if done == False and fourthClassTraining != "" and fourthClassTraining != None and fourthClassTraining != " ":
        day = day + "                0700: 4/c: " + fourthClassTraining
        if fourthClassLocation != "" and fourthClassLocation != None and fourthClassLocation != " ":
            day = day + "(" + fourthClassLocation + ")"
        day = day + """<br>
"""
else:
    if firstClassTraining != "" and firstClassTraining != None and firstClassTraining != " ":
        day = day + "                0800: 1/c: " + firstClassTraining
        if firstClassLocation != "" and firstClassLocation != None and firstClassLocation != " ":
            day = day + "(" + firstClassLocation + ")"
        day = day + """<br>
"""
    if secondClassTraining != "" and secondClassTraining != None and secondClassTraining != " ":
        day = day + "                0800: 2/c: " + secondClassTraining
        if secondClassLocation != "" and secondClassLocation != None and secondClassLocation != " ":
            day = day + "(" + secondClassLocation + ")"
        day = day + """<br>
"""
    if thirdClassTraining != "" and thirdClassTraining != None and thirdClassTraining != " ":
        day = day + "                0800: 3/c: " + thirdClassTraining
        if thirdClassLocation != "" and thirdClassLocation != None and thirdClassLocation != " ":
            day = day + "(" + thirdClassLocation + ")"
        day = day + """<br>
"""
    if fourthClassTraining != "" and fourthClassTraining != None and fourthClassTraining != " ":
        day = day + "                0800: 4/c: " + fourthClassTraining
        if fourthClassLocation != "" and fourthClassLocation != None and fourthClassLocation != " ":
            day = day + "(" + fourthClassLocation + ")"
        day = day + """<br>
"""
if dayOfWeek == "Mon" or dayOfWeek == "Wed":
    day = day + """                0800-1155: Morning Classes<br>
                1205: Afternoon Formation<br>
                1210: Family Style Lunch<br>
                1250-1540: Afternoon Classes<br>
                1600-1800: Sports Period<br>
                1700-1900: Buffet Dinner<br>
"""
elif dayOfWeek == "Tue" or dayOfWeek == "Thu":
    day = day + """                0800-1205: Morning Classes<br>
                1220: Afternoon Formation<br>
                1225: Family Style Lunch<br>
                1300-1540: Afternoon Classes<br>
                1600-1800: Sports Period<br>
                1700-1900: Buffet Dinner<br>
"""
elif dayOfWeek == "Fri":
    day = day + """                0800-1155: Morning Classes<br>
                1205: Afternoon Formation<br>
                1210: Family Style Lunch<br>
                1250-1540: Afternoon Classes<br>
                1600-1800: Sports Period<br>
                1800-1900: Buffet Dinner<br>
"""
elif dayOfWeek == "Sat":
    day = day + """                1200: Afternoon Formation<br>
                1200-1300: Buffet Style Lunch<br>
                1300: Restricted Cadet Formation<br>
                1600: Restricted Cadet Formation<br>
                1700-1800: Buffet Dinner<br>
"""
elif dayOfWeek == "Sun":
    day = day + """                1200-1300: Buffet Style Lunch<br>
                1300: Restricted Cadet Formation<br>
                1600: Restricted Cadet Formation<br>
                1700-1800: Buffet Dinner<br>
"""
if eveningTraining != "" and eveningTraining != None and eveningTraining != " ":
    day = day + "                1900: " + eveningTraining
    if eveningLocation != "" and eveningLocation != None:
        day = day + "(" + eveningLocation + ")"
    day = day + """<br>
"""
day = day + """                1930: Restricted Cadet Formation<br>
                2200: Taps/ Restricted Cadet Formation<br>
            </font>
        </div>"""
weatherText = """
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
                WX FORECAST: <b>""" + weather['day']['text'] + """</b><br>
                High: """ + weather['high'] + """ F<br>
                Low: """ + weather['low'] + """ F<br>
                """ + weather['day']['chance_precip'] + """% chance of precipitation<br>
                Sunrise: """ + datetime.time(hour=int(weather['sunrise'][0:1]), minute=int(weather['sunrise'][2:4])).strftime("%H%M") + """<br>
                Sunset: """ + datetime.time(hour=12+int(weather['sunset'][0:1]), minute=int(weather['sunset'][2:4])).strftime("%H%M") + """<br>
            </font>
        </div>"""
upcommingEventsText = """
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
                Upcoming events:<br>
                """
for event in upcomingEvents:
        upcommingEventsText += event
upcommingEventsText += """
            </font>
        </div>"""
footer = """
        <div style="font-size:13.3333px; margin-top:14pt; margin-bottom:14pt">
            <font face="Tahoma,sans-serif">
                CADET RECALL STATUS: <b>B-24</b>
            </font>
        </div>
    </body>
</html>"""
text = final + header + duty + day + weatherText + upcommingEventsText + footer
text = ''.join(i for i in text if ord(i)<128)
print text

me = "uscgapod@yahoo.com"
msg = MIMEText(text, 'html')
msg['Subject']= "POD " + tomorrow.strftime("%d%b%y")
msg['From'] = me
msg['To'] = "william.p.maxam@uscga.edu"

# Send the message via local SMTP server.
url = "smtp.mail.yahoo.com"
conn = smtplib.SMTP_SSL(url,465)
#conn.ehlo()
#conn.starttls()
#conn.ehlo()
user,password = (me,"P@ssword")
conn.login(user,password)
# sendmail function takes 3 arguments: sender's address, recipient's address
# and message to send - here it is sent as one string.
#conn.sendmail(me, you, msg.as_string())
conn.sendmail(me, "william.p.maxam@uscga.edu", msg.as_string())
print "Sent to william.p.maxam@uscga.edu"
conn.sendmail(me, "Richard.S.Lee@uscga.edu", msg.as_string())
print "Sent to Richard.S.Lee@uscga.edu"
conn.sendmail(me, "Silas.R.Garrett@uscga.edu", msg.as_string())
print "Sent to Silas.R.Garrett@uscga.edu"
conn.sendmail(me, "Matthew.S.Schoen@uscga.edu", msg.as_string())
print "Sent to Matthew.S.Schoen@uscga.edu"
conn.sendmail(me, "Jordan.T.Keller@uscga.edu", msg.as_string())
print "Sent to Jordan.T.Keller@uscga.edu"
conn.sendmail(me, "Michael.R.McCaslin@uscga.edu", msg.as_string())
print "Sent to Michael.R.McCaslin@uscga.edu"
conn.quit()
